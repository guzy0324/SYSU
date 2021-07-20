# -*- coding:utf-8 -*
import openpyxl
import nltk
from textblob.classifiers import NaiveBayesClassifier, DecisionTreeClassifier, MaxEntClassifier
from textblob import TextBlob
import re

def readxl():
    vine_data = list()
    data = list()
    for file in ('hair_dryer', 'microwave', 'pacifier'):
        book = openpyxl.load_workbook('../data/' + file + '.xlsx')
        sheet = book[file]
        
        for i in sheet.rows:
            rate = i[7].value
            vine = i[10].value == 'Y'
            purchase = i[11].value == 'Y'
            title = str(i[12].value)
            review = str(i[13].value)
            if type(rate) != int or not (vine or purchase) or not (rate == 1 or rate == 2 or rate == 5):
                continue
            if vine:
                vine_data.append((title + '. ' + review, 'pos' if rate == 5 else 'neg'))
            else:
                data.append((title + '. ' + review, 'pos' if rate == 5 else 'neg'))
    return vine_data, data

from random import shuffle
def train(vine_data, data):
    shuffle(data)
    print(len(vine_data))
    print(len(data))
    rate = 0.7
    boundary = int(rate * (len(vine_data) + len(data)))
    with open('../data/accurate.txt', 'a') as f:
        f.write('\ntotal size: ' + str(boundary) + '\n')
        cl2 = DecisionTreeClassifier(vine_data + data[:boundary - len(vine_data)])
        accurate2 = cl2.accuracy(data[boundary - len(vine_data):])
        print(accurate2)
    return cl2

def analyze(cl, number):
    for file in ('hair_dryer', 'microwave', 'pacifier'):
        with open('../data/' + number + file + '.txt', 'a') as f:
            book = openpyxl.load_workbook('../data/' + file + '.xlsx')
            sheet = book[file]
            
            for i in sheet.rows:
                title = str(i[12].value)
                review = str(i[13].value)
                prob_dist = cl.prob_classify(title + review)
                f.write(str(round(prob_dist.prob("pos"), 2)) + ' ')
                f.write(str(round(prob_dist.prob("neg"), 2)) + '\n')

def count(number):
    for file in ('hair_dryer', 'microwave', 'pacifier'):
        with open('../data/' + number + file + 'length.txt', 'a') as f:
            book = openpyxl.load_workbook('../data/' + file + '.xlsx')
            sheet = book[file]
            
            for i in sheet.rows:
                title = str(i[12].value)
                review = str(i[13].value)
                pat_letter = re.compile(r'[^a-zA-Z \']+')
                new_text = pat_letter.sub(' ', title + '. ' + review).strip().lower()
                f.write(str(len(nltk.word_tokenize(new_text))) + '\n')
        